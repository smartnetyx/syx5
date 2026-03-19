"""
TERP Contract Management Graph Visualization
Streamlit Application for Contract Network Analysis
"""

import streamlit as st
import pandas as pd
import json
import plotly.graph_objects as go
import plotly.express as px
import networkx as nx
from pathlib import Path
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import warnings
warnings.filterwarnings('ignore')

# Page configuration
st.set_page_config(
    page_title="TERP Vertragsmanagement",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── PASSWORD PROTECTION ──
def get_app_password():
    """Get password from st.secrets (lokal/Streamlit Cloud) or env var (Azure)."""
    try:
        return st.secrets["app_password"]
    except (FileNotFoundError, KeyError):
        return os.environ.get("APP_PASSWORD", "")

def check_password():
    """Password gate: URL token (for D365 iframe) or manual login."""
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    app_password = get_app_password()

    # Auto-login via URL token (?token=...) for iframe embedding in Dynamics 365
    try:
        token = st.query_params.get("token")
    except AttributeError:
        token = st.experimental_get_query_params().get("token", [None])[0]
    if token and app_password and token == app_password:
        st.session_state.authenticated = True
        return True

    # Manual login form
    st.markdown("""
    <h2 style='text-align: center; color: #2C3E50; margin-top: 100px;'>
        TERP Vertragsmanagement
    </h2>
    <p style='text-align: center; color: #7F8C8D;'>
        Bitte Passwort eingeben, um fortzufahren.
    </p>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        password = st.text_input("Passwort", type="password", key="password_input")
        if st.button("Anmelden", use_container_width=True):
            if app_password and password == app_password:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Falsches Passwort.")
    return False

if not check_password():
    st.stop()

# Custom CSS — TERP Model-driven App Design
st.markdown("""<style>
/* Global font and background */
html, body, [class*="css"] {
    font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif !important;
    color: #333333 !important;
}
/* Header: visibility hidden statt display none — Kinder bleiben im Rendering-Tree */
header[data-testid="stHeader"] {
    visibility: hidden !important;
    height: 0 !important;
    min-height: 0 !important;
    padding: 0 !important;
    margin: 0 !important;
    overflow: visible !important;
    background: transparent !important;
    border: none !important;
}

/* Sidebar expand button — sichtbar aus dem hidden Header herauslösen */
[data-testid="stExpandSidebarButton"] {
    position: fixed !important;
    display: flex !important;
    visibility: visible !important;
    top: 6px !important;
    left: 6px !important;
    z-index: 999999 !important;
    width: 36px !important;
    height: 36px !important;
    align-items: center !important;
    justify-content: center !important;
    background-color: #742774 !important;
    border-radius: 4px !important;
    box-shadow: 0 2px 6px rgba(0,0,0,0.3) !important;
    cursor: pointer !important;
    opacity: 1 !important;
}
[data-testid="stExpandSidebarButton"] * {
    visibility: visible !important;
    color: white !important;
    fill: white !important;
    stroke: white !important;
}
[data-testid="stExpandSidebarButton"]:hover {
    background-color: #5a1f5a !important;
}

/* Sidebar collapse button — gleiche Größe und Farbe wie expand */
[data-testid="stSidebarCollapseButton"] {
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    width: 36px !important;
    height: 36px !important;
    background-color: #742774 !important;
    border-radius: 4px !important;
    cursor: pointer !important;
}
[data-testid="stSidebarCollapseButton"] *,
[data-testid="stSidebarCollapseButton"] svg,
[data-testid="stSidebarCollapseButton"] svg * {
    color: white !important;
    fill: white !important;
    stroke: white !important;
}
[data-testid="stSidebarCollapseButton"]:hover {
    background-color: #5a1f5a !important;
}
.stApp { background-color: #E8E8E8 !important; }
.block-container {
    padding-top: 0 !important;
    background-color: #FFFFFF !important;
    border: 1px solid #D0D0D0 !important;
    border-radius: 4px !important;
    margin: 0.5rem !important;
}

/* TERP Header bar */
.terp-header {
    background-color: #742774;
    color: #FFFFFF;
    padding: 0 24px;
    height: 48px;
    display: flex;
    align-items: center;
    font-size: 16px;
    font-weight: 600;
    font-family: 'Segoe UI', sans-serif;
    margin: -1rem -1rem 0.5rem -1rem;
    width: calc(100% + 2rem);
}

/* Tab styling — MDA-style: bold + blue underline */
button[data-baseweb="tab"] > div > p {
    font-size: 15px !important;
    font-family: 'Segoe UI', sans-serif !important;
}
button[data-baseweb="tab"] {
    margin-right: 4px !important;
    border-radius: 0 !important;
    background-color: transparent !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    background-color: transparent !important;
    border-bottom: 3px solid #0f6cbd !important;
    border-radius: 0 !important;
}
button[data-baseweb="tab"][aria-selected="true"] > div > p {
    color: #333333 !important;
    font-weight: 700 !important;
}
button[data-baseweb="tab"][aria-selected="false"] {
    border-bottom: 3px solid transparent !important;
}
button[data-baseweb="tab"][aria-selected="false"] > div > p {
    color: #333333 !important;
    font-weight: 400 !important;
}

/* Kill Streamlit default tab highlight/border bar (prevents double underline) */
div[data-baseweb="tab-highlight"] { display: none !important; }
div[data-baseweb="tab-border"] { display: none !important; }

/*
 * ╔══════════════════════════════════════════════════════════╗
 * ║  SIDEBAR LAYOUT — Anpassbare Parameter                  ║
 * ║                                                          ║
 * ║  sidebar-bg:       Hintergrund der Sidebar (#F8F8F8)    ║
 * ║  sidebar-border:   Rechter Rand (#D0D0D0)               ║
 * ║  sidebar-padding:  Innenabstand links/rechts (8px)       ║
 * ║  sidebar-top:      Innenabstand oben (12px)              ║
 * ║                                                          ║
 * ║  Zum Anpassen: Werte unten ändern und App neu laden.    ║
 * ╚══════════════════════════════════════════════════════════╝
 */
[data-testid="stSidebar"] {
    background-color: #F8F8F8 !important;              /* sidebar-bg */
    border-right: 1px solid #D0D0D0 !important;        /* sidebar-border */
}
[data-testid="stSidebar"] [data-testid="stSidebarCollapsedControl"],
[data-testid="stSidebar"]::after {
    border-right-color: #D0D0D0 !important;            /* sidebar-border */
}
[data-testid="stSidebar"] > div,
[data-testid="stSidebar"] > div:first-child,
[data-testid="stSidebar"] [data-testid="stSidebarContent"],
[data-testid="stSidebar"] section[data-testid="stSidebarContent"] > div {
    padding-left: 8px !important;                      /* sidebar-padding */
    padding-right: 8px !important;                     /* sidebar-padding */
    padding-top: 12px !important;                      /* sidebar-top */
}
/* Dropdowns: Mindestbreite entfernen, volle Breite erzwingen */
[data-testid="stSidebar"] [data-baseweb="select"],
[data-testid="stSidebar"] [data-baseweb="popover"],
[data-testid="stSidebar"] .stMultiSelect,
[data-testid="stSidebar"] .stSelectbox {
    min-width: 0 !important;
    width: 100% !important;
}
[data-testid="stSidebar"] [data-baseweb="select"] > div {
    min-width: 0 !important;
}
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] {
    color: #333333 !important;
}
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] h4 {
    color: #333333 !important;
    font-weight: 600 !important;
}
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h2 .icon,
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h4 .icon { display: none !important; }

/* Dropdown/Select styling */
[data-baseweb="select"] > div {
    background-color: #F5F5F5 !important;
    border: 1px solid #D0D0D0 !important;
    border-radius: 4px !important;
    color: #333333 !important;
}

/* Sidebar — reduce spacing around dividers and top padding */
[data-testid="stSidebar"] > div:first-child {
    padding-top: 0.5rem !important;
}
[data-testid="stSidebar"] hr {
    margin-top: 0.3rem !important;
    margin-bottom: 0.3rem !important;
}
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h4 {
    margin-top: 0.2rem !important;
    margin-bottom: 0.2rem !important;
}

/* Slider: TERP-Lila statt Rot — breite Selektoren fuer alle Streamlit-Versionen */
[data-testid="stSlider"] [role="slider"],
[data-testid="stSlider"] [data-testid="stThumbValue"],
div[data-baseweb="slider"] [role="slider"] {
    background-color: #742774 !important;
}
div[data-baseweb="slider"] div[role="progressbar"] > div,
div[data-baseweb="slider"] div[role="progressbar"],
[data-testid="stSlider"] div[data-baseweb="slider"] div[style*="background"] {
    background-color: #742774 !important;
}

/* Radio: TERP-Lila statt Rot */
div[data-baseweb="radio"] div[aria-checked="true"] > div:first-child {
    background-color: #742774 !important;
    border-color: #742774 !important;
}

/* Filter-Reset-Button in Sidebar: Dropdown-Stil (grau, dunkler Rahmen) */
[data-testid="stSidebar"] .stButton > button,
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] > button {
    background-color: #F5F5F5 !important;
    color: #333333 !important;
    border: 1px solid #D0D0D0 !important;
    border-radius: 4px !important;
    font-size: 13px !important;
}
[data-testid="stSidebar"] .stButton > button:hover,
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] > button:hover {
    background-color: #E8E8E8 !important;
    color: #333333 !important;
    border: 1px solid #AAAAAA !important;
}

/* Metric styling */
[data-testid="stMetricValue"] {
    color: #333333 !important;
    font-weight: 600 !important;
}
[data-testid="stMetricLabel"] {
    color: #666666 !important;
}
</style>""", unsafe_allow_html=True)

# TERP Header bar
st.markdown('<div class="terp-header">TERP Vertragsmanagement (v1.4.8)</div>', unsafe_allow_html=True)

# Color scheme for node types — TERP pastel palette
NODE_COLORS = {
    'Vertrag':  {'background': '#FABBC3', 'border': '#E8949F', 'highlight': {'background': '#FDD5DA', 'border': '#E8949F'}, 'hover': {'background': '#FDD5DA', 'border': '#E8949F'}},
    'Firma':    {'background': '#CAEEE9', 'border': '#9DD5CC', 'highlight': {'background': '#DFF5F2', 'border': '#9DD5CC'}, 'hover': {'background': '#DFF5F2', 'border': '#9DD5CC'}},
    'Person':   {'background': '#FFEBC0', 'border': '#E8CF97', 'highlight': {'background': '#FFF3DA', 'border': '#E8CF97'}, 'hover': {'background': '#FFF3DA', 'border': '#E8CF97'}},
    'Anlage':   {'background': '#D4E7F6', 'border': '#A8CCE6', 'highlight': {'background': '#E5F0FA', 'border': '#A8CCE6'}, 'hover': {'background': '#E5F0FA', 'border': '#A8CCE6'}},
    'Projekt':  {'background': '#D8D8EE', 'border': '#B3B3D6', 'highlight': {'background': '#E8E8F5', 'border': '#B3B3D6'}, 'hover': {'background': '#E8E8F5', 'border': '#B3B3D6'}},
    'Dokument': {'background': '#E5E5E5', 'border': '#C0C0C0', 'highlight': {'background': '#F0F0F0', 'border': '#C0C0C0'}, 'hover': {'background': '#F0F0F0', 'border': '#C0C0C0'}},
    'MaLo':     {'background': '#E6E0F0', 'border': '#C5BAD9', 'highlight': {'background': '#F0ECF5', 'border': '#C5BAD9'}, 'hover': {'background': '#F0ECF5', 'border': '#C5BAD9'}},
}

# Simple colors for legend display
NODE_COLORS_SIMPLE = {
    'Vertrag':  '#FABBC3',
    'Firma':    '#CAEEE9',
    'Person':   '#FFEBC0',
    'Anlage':   '#D4E7F6',
    'Projekt':  '#D8D8EE',
    'Dokument': '#E5E5E5',
    'MaLo':     '#E6E0F0',
}

# Load data
@st.cache_resource
def load_data():
    """Load vertices, edges, and contracts metadata"""
    data_dir = next(p for p in [Path(__file__).parent / 'data', Path(__file__).parent.parent / 'data'] if p.exists())

    with open(data_dir / 'vertices.json', 'r', encoding='utf-8') as f:
        vertices = json.load(f)

    with open(data_dir / 'edges.json', 'r', encoding='utf-8') as f:
        edges = json.load(f)

    with open(data_dir / 'contracts_metadata.json', 'r', encoding='utf-8') as f:
        contracts = json.load(f)

    return vertices, edges, contracts

@st.cache_data
def build_network_graph(vertices, edges):
    """Build NetworkX graph from vertices and edges"""
    G = nx.DiGraph()

    # Add nodes with attributes
    for vertex in vertices:
        G.add_node(
            vertex['id'],
            label=vertex.get('label', 'Unbekannt'),
            name=vertex.get('name', vertex['id']),
            **{k: v for k, v in vertex.items() if k not in ['id', 'label', 'name']}
        )

    # Add edges with attributes
    for edge in edges:
        G.add_edge(
            edge['src'],
            edge['dst'],
            relationship=edge.get('relationship', 'UNKNOWN'),
            **{k: v for k, v in edge.items() if k not in ['src', 'dst', 'relationship']}
        )

    return G

@st.cache_data
def get_firmen_list(vertices):
    """Get unique list of Firmen"""
    firmen = [v for v in vertices if v.get('label') == 'Firma']
    return sorted([f.get('name', f['id']) for f in firmen])

@st.cache_data
def get_personen_list(vertices):
    """Get unique list of Personen"""
    personen = [v for v in vertices if v.get('label') == 'Person']
    return sorted([p.get('name', p['id']) for p in personen])

@st.cache_data
def get_vertragsart_list(contracts):
    """Get unique list of Vertragsart"""
    return sorted(set(c.get('vertragsart', 'Unbekannt') for c in contracts if c.get('vertragsart')))

@st.cache_data
def get_projekte_list(vertices):
    """Get unique list of Projekte"""
    projekte = [v for v in vertices if v.get('label') == 'Projekt']
    return sorted([p.get('name', p['id']) for p in projekte])

def filter_contracts(contracts, vertices, edges, firmen_filter, projekt_filter, vertragsart_filter, personen_filter,
                     standort_filter=None, rolle_filter=None, nachtrag_filter=None, seiten_range=None):
    """Filter contracts based on all filters"""
    filtered = list(contracts)

    if vertragsart_filter:
        filtered = [c for c in filtered if c.get('vertragsart') in vertragsart_filter]

    if firmen_filter:
        filtered = [c for c in filtered
                   if any(p.get('name') in firmen_filter for p in c.get('parteien', []))]

    if personen_filter:
        filtered = [c for c in filtered
                   if any(p.get('name') in personen_filter for p in c.get('parteien', []))]

    if projekt_filter:
        projekt_nodes = [v['id'] for v in vertices
                        if v.get('label') == 'Projekt' and v.get('name') in projekt_filter]
        contract_ids_with_projekt = set()
        for edge in edges:
            if edge['dst'] in projekt_nodes and 'vertrag:' in edge['src']:
                contract_ids_with_projekt.add(edge['src'])
            if edge['src'] in projekt_nodes and 'vertrag:' in edge['dst']:
                contract_ids_with_projekt.add(edge['dst'])
        filtered = [c for c in filtered if c.get('vertrags_id') in contract_ids_with_projekt]

    if standort_filter:
        filtered = [c for c in filtered
                   if any(s in standort_filter
                         for s in (c.get('standorte', []) if isinstance(c.get('standorte'), list) else [c.get('standorte', '')]))]

    if rolle_filter:
        filtered_by_rolle = []
        for c in filtered:
            c_id = c.get('vertrags_id')
            for edge in edges:
                if ((edge['dst'] == c_id or edge['src'] == c_id)
                    and edge.get('rolle') in rolle_filter):
                    filtered_by_rolle.append(c)
                    break
        filtered = filtered_by_rolle

    if nachtrag_filter == 'Nur Nachträge':
        filtered = [c for c in filtered if c.get('ist_nachtrag')]
    elif nachtrag_filter == 'Nur Originale':
        filtered = [c for c in filtered if not c.get('ist_nachtrag')]

    if seiten_range:
        filtered = [c for c in filtered
                   if seiten_range[0] <= (c.get('seitenzahl') or 0) <= seiten_range[1]]

    return filtered

def filter_graph_by_criteria(G, vertices, edges, firmen_filter, vertragsart_filter, projekt_filter, personen_filter=None):
    """Filter graph based on selected criteria"""
    contracts_filtered = [v['id'] for v in vertices
                         if v.get('label') == 'Vertrag']

    if vertragsart_filter:
        contracts_filtered = [c for c in contracts_filtered
                            if any(v['id'] == c and v.get('vertragsart') in vertragsart_filter
                                   for v in vertices)]

    if firmen_filter:
        firma_nodes = [v['id'] for v in vertices
                      if v.get('label') == 'Firma' and v.get('name') in firmen_filter]
        contracts_with_firmen = set()
        for contract in contracts_filtered:
            for edge in edges:
                if ((edge['src'] in firma_nodes and edge['dst'] == contract) or
                    (edge['src'] == contract and edge['dst'] in firma_nodes)):
                    contracts_with_firmen.add(contract)
        contracts_filtered = list(contracts_with_firmen)

    if personen_filter:
        person_nodes = [v['id'] for v in vertices
                       if v.get('label') == 'Person' and v.get('name') in personen_filter]
        contracts_with_personen = set()
        for contract in contracts_filtered:
            for edge in edges:
                if ((edge['src'] in person_nodes and edge['dst'] == contract) or
                    (edge['src'] == contract and edge['dst'] in person_nodes)):
                    contracts_with_personen.add(contract)
        contracts_filtered = list(contracts_with_personen)

    if projekt_filter:
        projekt_nodes = [v['id'] for v in vertices
                        if v.get('label') == 'Projekt' and v.get('name') in projekt_filter]
        contracts_with_projekten = set()
        for contract in contracts_filtered:
            for edge in edges:
                if ((edge['src'] == contract and edge['dst'] in projekt_nodes) or
                    (edge['src'] in projekt_nodes and edge['dst'] == contract)):
                    contracts_with_projekten.add(contract)
        contracts_filtered = list(contracts_with_projekten)

    relevant_nodes = set(contracts_filtered)
    for contract in contracts_filtered:
        for edge in edges:
            if edge['src'] == contract:
                relevant_nodes.add(edge['dst'])
            if edge['dst'] == contract:
                relevant_nodes.add(edge['src'])

    subgraph_vertices = [v for v in vertices if v['id'] in relevant_nodes]
    subgraph_edges = [e for e in edges
                     if e['src'] in relevant_nodes and e['dst'] in relevant_nodes]

    subG = nx.DiGraph()
    for vertex in subgraph_vertices:
        subG.add_node(
            vertex['id'],
            **{k: v for k, v in vertex.items() if k != 'id'}
        )

    for edge in subgraph_edges:
        subG.add_edge(edge['src'], edge['dst'],
                     **{k: v for k, v in edge.items() if k not in ['src', 'dst']})

    return subG

def create_pyvis_graph(G, spring_length=180, gravity=-8000, central_gravity=0.10):
    """Create interactive pyvis visualization"""
    import pyvis.network as net

    g = net.Network(
        height='900px',
        width='100%',
        directed=True,
        notebook=False,
        layout=None
    )

    # German labels for metadata keys
    KEY_LABELS = {
        'vertragsart': 'Vertragsart', 'status': 'Status', 'kategorie': 'Kategorie',
        'seitenzahl': 'Seiten', 'standorte': 'Standorte', 'vertragswert': 'Vertragswert',
        'datum_gefunden': 'Datum', 'ist_nachtrag': 'Nachtrag', 'rolle': 'Rolle',
        'typ': 'Typ', 'ort': 'Ort', 'standort': 'Standort', 'dateipfad': 'Dateipfad',
        'dokument_typ': 'Dokument-Typ', 'beziehung': 'Beziehung',
        'anteil_prozent': 'Anteil (%)', 'produkt': 'Produkt', 'vertrag': 'Vertrag',
    }

    # Pre-build contract metadata lookup for rich tooltips
    contract_meta = {}
    try:
        data_dir = next(p for p in [Path(__file__).parent / 'data', Path(__file__).parent.parent / 'data'] if p.exists())
        with open(data_dir / 'contracts_metadata.json', 'r', encoding='utf-8') as f:
            for cm in json.load(f):
                contract_meta[cm.get('vertrags_id')] = cm
    except:
        pass

    for node in G.nodes():
        node_data = G.nodes[node]
        label = node_data.get('label', 'Unbekannt')
        color = NODE_COLORS.get(label, {'background': '#E5E5E5', 'border': '#C0C0C0'})
        name = node_data.get('name', node)[:50]

        # Build rich tooltip based on node type
        tooltip_lines = [f"[{label}] {node_data.get('name', node)}"]

        if label == 'Vertrag':
            meta = contract_meta.get(node, {})
            if node_data.get('vertragsart'):
                tooltip_lines.append(f"Vertragsart: {node_data['vertragsart']}")
            if node_data.get('kategorie'):
                tooltip_lines.append(f"Kategorie: {node_data['kategorie']}")
            if node_data.get('vertragstitel'):
                tooltip_lines.append(f"Titel: {node_data['vertragstitel']}")
            if meta.get('vertragsgegenstand'):
                tooltip_lines.append(f"Gegenstand: {meta['vertragsgegenstand']}")
            # Parteien
            parteien = meta.get('parteien', [])
            if parteien:
                for p in parteien:
                    rolle = p.get('rolle', '')
                    name_p = p.get('name', '')
                    tooltip_lines.append(f"  {rolle}: {name_p}" if rolle else f"  Partei: {name_p}")
            # Datum
            datum = meta.get('datum', {})
            if datum.get('abschluss'):
                tooltip_lines.append(f"Abschluss: {datum['abschluss']}")
            if datum.get('beginn'):
                tooltip_lines.append(f"Beginn: {datum['beginn']}")
            if datum.get('ende'):
                tooltip_lines.append(f"Ende: {datum['ende']}")
            # Laufzeit
            if meta.get('laufzeit'):
                tooltip_lines.append(f"Laufzeit: {meta['laufzeit']}")
            if meta.get('auto_verlaengerung'):
                tooltip_lines.append(f"Verlängerung: {meta['auto_verlaengerung']}")
            # Kündigung
            if meta.get('kuendigungsfrist'):
                tooltip_lines.append(f"Kündigung: {meta['kuendigungsfrist']}")
            if meta.get('kuendigung_form'):
                tooltip_lines.append(f"Kündigungsform: {meta['kuendigung_form']}")
            if meta.get('kuendigung_ausserordentlich'):
                tooltip_lines.append(f"Außerordentlich: {meta['kuendigung_ausserordentlich']}")
            # Finanzen
            fin = meta.get('finanzen', {})
            if fin.get('verguetung'):
                tooltip_lines.append(f"Vergütung: {fin['verguetung']}")
            if fin.get('verguetungsbasis'):
                tooltip_lines.append(f"Vergütungsbasis: {fin['verguetungsbasis']}")
            if fin.get('vertragswert_jaehrlich'):
                tooltip_lines.append(f"Vertragswert p.a.: {fin['vertragswert_jaehrlich']}")
            if fin.get('arbeitspreis_ct_kwh'):
                tooltip_lines.append(f"Arbeitspreis: {fin['arbeitspreis_ct_kwh']} ct/kWh")
            if fin.get('strompreis_ct_kwh'):
                tooltip_lines.append(f"Strompreis: {fin['strompreis_ct_kwh']} ct/kWh")
            if fin.get('kaltmiete_eur'):
                tooltip_lines.append(f"Kaltmiete: {fin['kaltmiete_eur']} EUR")
            if fin.get('nebenkosten_eur'):
                tooltip_lines.append(f"Nebenkosten: {fin['nebenkosten_eur']} EUR")
            if fin.get('gesamtmiete_eur'):
                tooltip_lines.append(f"Gesamtmiete: {fin['gesamtmiete_eur']} EUR")
            if fin.get('zahlungsrhythmus'):
                tooltip_lines.append(f"Zahlungsrhythmus: {fin['zahlungsrhythmus']}")
            if fin.get('waehrung') and fin['waehrung'] != 'EUR':
                tooltip_lines.append(f"Währung: {fin['waehrung']}")
            # Energie
            ene = meta.get('energie', {})
            if ene.get('pv_leistung_kwp'):
                tooltip_lines.append(f"PV: {ene['pv_leistung_kwp']} kWp")
            if ene.get('anschlussleistung_kw'):
                tooltip_lines.append(f"Anschlussleistung: {ene['anschlussleistung_kw']} kW")
            if ene.get('malo_leistung_mw'):
                tooltip_lines.append(f"MaLo-Leistung: {ene['malo_leistung_mw']} MW")
            malo = ene.get('malo_ids', [])
            if malo:
                tooltip_lines.append(f"MaLo-IDs: {', '.join(str(m) for m in malo)}")
            # Standort
            standorte = meta.get('standorte', [])
            if standorte and any(standorte):
                tooltip_lines.append(f"Standort: {', '.join(s for s in standorte if s)}")
            if meta.get('gerichtsstand'):
                tooltip_lines.append(f"Gerichtsstand: {meta['gerichtsstand']}")
            # Nachtrag / Status
            if node_data.get('ist_nachtrag'):
                tooltip_lines.append("Status: NACHTRAG")
            tooltip_lines.append(f"Seiten: {node_data.get('seitenzahl', '?')}")
            if meta.get('dateipfad'):
                tooltip_lines.append(f"Datei: {meta['dateipfad']}")

        elif label == 'Firma':
            if node_data.get('rechtsform'):
                tooltip_lines.append(f"Rechtsform: {node_data['rechtsform']}")
            if node_data.get('adresse'):
                tooltip_lines.append(f"Adresse: {node_data['adresse']}")
            n_contracts = sum(1 for _, dst in G.edges(node) if G.nodes.get(dst, {}).get('label') == 'Vertrag')
            n_contracts += sum(1 for src, _ in G.in_edges(node) if G.nodes.get(src, {}).get('label') == 'Vertrag')
            if n_contracts:
                tooltip_lines.append(f"Verträge: {n_contracts}")

        elif label == 'Person':
            if node_data.get('rechtsform'):
                tooltip_lines.append(f"Rechtsform: {node_data['rechtsform']}")
            if node_data.get('adresse'):
                tooltip_lines.append(f"Adresse: {node_data['adresse']}")
            n_contracts = sum(1 for _, dst in G.edges(node) if G.nodes.get(dst, {}).get('label') == 'Vertrag')
            n_contracts += sum(1 for src, _ in G.in_edges(node) if G.nodes.get(src, {}).get('label') == 'Vertrag')
            if n_contracts:
                tooltip_lines.append(f"Verträge: {n_contracts}")

        elif label == 'Anlage':
            if node_data.get('typ'):
                tooltip_lines.append(f"Typ: {node_data['typ']}")
            if node_data.get('leistung_kwp'):
                tooltip_lines.append(f"Leistung: {node_data['leistung_kwp']} kWp")
            if node_data.get('standort'):
                tooltip_lines.append(f"Standort: {node_data['standort']}")

        elif label == 'MaLo':
            if node_data.get('malo_leistung_mw'):
                tooltip_lines.append(f"Leistung: {node_data['malo_leistung_mw']} MW")

        elif label == 'Dokument':
            if node_data.get('dateipfad'):
                tooltip_lines.append(f"Pfad: {node_data['dateipfad']}")
            if node_data.get('seitenzahl'):
                tooltip_lines.append(f"Seiten: {node_data['seitenzahl']}")

        elif label == 'Projekt':
            n_contracts = sum(1 for src, _ in G.in_edges(node) if G.nodes.get(src, {}).get('label') == 'Vertrag')
            if n_contracts:
                tooltip_lines.append(f"Verträge: {n_contracts}")

        # Alle verbleibenden Vertex-Attribute anzeigen (Catch-all)
        shown_keys = {'id', 'label', 'name', 'vertragsart', 'kategorie', 'vertragstitel',
                       'status', 'standorte', 'seitenzahl', 'datum_gefunden', 'vertragswert',
                       'datum_ende', 'datum_ende_jahr', 'ist_nachtrag', 'rechtsform', 'adresse',
                       'typ', 'leistung_kwp', 'standort', 'malo_leistung_mw', 'dateipfad'}
        for k, v in node_data.items():
            if k not in shown_keys and v:
                display_key = KEY_LABELS.get(k, k)
                tooltip_lines.append(f"{display_key}: {v}")

        title = "\n".join(tooltip_lines)

        g.add_node(
            node,
            label=name,
            title=title,
            color=color,
            size=28 if label == 'Vertrag' else 22,
            borderWidth=2,
            borderWidthSelected=3,
            shadow={'enabled': True, 'color': 'rgba(0,0,0,0.2)', 'size': 8, 'x': 3, 'y': 3},
            font={'size': 9, 'face': 'Segoe UI, Arial, sans-serif'}
        )

    for src, dst in G.edges():
        edge_data = G.edges[src, dst]
        relationship = edge_data.get('relationship', 'UNKNOWN')

        # Rich edge tooltip
        src_name = G.nodes.get(src, {}).get('name', src)
        dst_name = G.nodes.get(dst, {}).get('name', dst)
        tooltip_lines = [relationship]
        tooltip_lines.append(f"Von: {src_name}")
        tooltip_lines.append(f"An: {dst_name}")
        rolle = edge_data.get('rolle')
        if rolle:
            tooltip_lines.append(f"Rolle: {rolle}")
        if edge_data.get('dokument_typ'):
            tooltip_lines.append(f"Dokument-Typ: {edge_data['dokument_typ']}")
        for k, v in edge_data.items():
            if k not in ('relationship', 'rolle', 'dokument_typ') and v:
                display_key = KEY_LABELS.get(k, k)
                tooltip_lines.append(f"{display_key}: {v}")
        edge_title = "\n".join(tooltip_lines)

        # Show rolle on edge label if present
        edge_label = f"{relationship}\n({rolle})" if rolle else relationship

        g.add_edge(src, dst, label=edge_label,
                  title=edge_title, font={'size': 8},
                  color={'color': '#B0B0B0', 'highlight': '#742774', 'hover': '#742774'},
                  smooth={'type': 'curvedCW', 'roundness': 0.1})

    # Physics: configurable via parameters
    g.set_options(json.dumps({
        "physics": {
            "barnesHut": {
                "gravitationalConstant": gravity,
                "centralGravity": central_gravity,
                "springLength": spring_length,
                "springConstant": 0.02,
                "damping": 0.20,
                "avoidOverlap": 0.5
            },
            "minVelocity": 0.75
        }
    }))

    return g

def create_ego_pyvis_graph(subgraph):
    """Create pyvis graph optimized for ego-graph (shorter edges, compact)"""
    # Convert subgraph view to a regular DiGraph to avoid 'dict' attribute errors
    G_copy = nx.DiGraph()
    for node in subgraph.nodes():
        G_copy.add_node(node, **dict(subgraph.nodes[node]))
    for src, dst in subgraph.edges():
        G_copy.add_edge(src, dst, **dict(subgraph.edges[src, dst]))

    # Pass ego-specific physics params directly to avoid double set_options call
    g = create_pyvis_graph(G_copy, spring_length=100, gravity=-3000, central_gravity=0.30)
    return g

def create_contracts_table(contracts_list):
    """Create comprehensive contracts table with extended fields"""
    df = pd.DataFrame()
    df['Vertrags-ID'] = [c.get('vertrags_id', '') for c in contracts_list]
    df['Vertragsart'] = [c.get('vertragsart', '') for c in contracts_list]
    df['Titel'] = [c.get('vertragstitel', '') or '' for c in contracts_list]
    df['Parteien'] = [', '.join([p.get('name', '') for p in c.get('parteien', [])])
                     for c in contracts_list]
    df['Rollen'] = [', '.join([p.get('rolle', '') for p in c.get('parteien', []) if p.get('rolle')])
                   for c in contracts_list]
    df['Datum Abschluss'] = [c.get('datum', {}).get('abschluss', '') or '' for c in contracts_list]
    df['Datum Beginn'] = [c.get('datum', {}).get('beginn', '') or '' for c in contracts_list]
    df['Datum Ende'] = [c.get('datum', {}).get('ende', '') or '' for c in contracts_list]
    df['Laufzeit'] = [c.get('laufzeit', '') or '' for c in contracts_list]
    df['Kündigung'] = [c.get('kuendigungsfrist', '') or '' for c in contracts_list]
    df['Standort'] = [', '.join(c.get('standorte', [])) if isinstance(c.get('standorte'), list) else c.get('standorte', '') for c in contracts_list]
    df['Vergütung'] = [c.get('finanzen', {}).get('verguetung', '') or '' for c in contracts_list]
    df['Vertragswert p.a.'] = [c.get('finanzen', {}).get('vertragswert_jaehrlich', '') or '' for c in contracts_list]
    df['PV kWp'] = [c.get('energie', {}).get('pv_leistung_kwp', '') or '' for c in contracts_list]
    df['MaLo-IDs'] = [', '.join(c.get('energie', {}).get('malo_ids', []) or []) for c in contracts_list]
    df['Status'] = ['Nachtrag' if c.get('ist_nachtrag') else 'Original' for c in contracts_list]
    df['Seiten'] = [c.get('seitenzahl', '') for c in contracts_list]
    df['Gerichtsstand'] = [c.get('gerichtsstand', '') or '' for c in contracts_list]

    return df

def get_contracts_by_firma(contracts, selected_firmen):
    """Get contracts grouped by company"""
    result = {}
    for firma in selected_firmen:
        firma_contracts = []
        for c in contracts:
            for party in c.get('parteien', []):
                if party.get('name') == firma:
                    firma_contracts.append(c)
                    break
        result[firma] = firma_contracts
    return result

def get_contracts_by_vertragsart(contracts, selected_arten):
    """Get contracts grouped by type"""
    result = {}
    for art in selected_arten:
        result[art] = [c for c in contracts if c.get('vertragsart') == art]
    return result

def get_contracts_by_projekt(vertices, edges, contracts):
    """Get contracts grouped by project"""
    result = {}
    projekte = [v for v in vertices if v.get('label') == 'Projekt']

    for projekt in projekte:
        projekt_name = projekt.get('name')
        projekt_contracts = []

        for edge in edges:
            if edge['dst'] == projekt['id'] and 'vertrag:' in edge['src']:
                vertrag_id = edge['src']
                matching = [c for c in contracts if c.get('vertrags_id') == vertrag_id]
                if matching:
                    projekt_contracts.append(matching[0])

        if projekt_contracts:
            result[projekt_name] = projekt_contracts

    return result

def get_parteien_ubersicht(contracts):
    """Get overview of all parties and their contracts"""
    parteien_map = {}

    for c in contracts:
        for party in c.get('parteien', []):
            party_name = party.get('name')
            if party_name not in parteien_map:
                parteien_map[party_name] = {
                    'Typ': party.get('typ'),
                    'Anzahl Verträge': 0,
                    'Verträge': []
                }
            parteien_map[party_name]['Anzahl Verträge'] += 1
            parteien_map[party_name]['Verträge'].append(c.get('dateiname', ''))

    return parteien_map

def get_finanz_ubersicht(contracts):
    """Get financial overview"""
    df = pd.DataFrame()
    df['Vertrags-ID'] = [c.get('vertrags_id', '') for c in contracts]
    df['Dateiname'] = [c.get('dateiname', '') for c in contracts]
    df['Vertragsart'] = [c.get('vertragsart', '') for c in contracts]

    finanzen_list = [c.get('finanzen', {}) for c in contracts]
    non_empty = [f for f in finanzen_list if f]
    if non_empty:
        for key in set().union(*[set(f.keys()) for f in non_empty]):
            df[key] = [f.get(key, '') for f in finanzen_list]

    return df

def get_waermeabnehmer_ubersicht(edges, vertices):
    """Get overview of heat supply customers"""
    waermeabnehmer = []

    for edge in edges:
        if edge.get('relationship') == 'LIEFERT_WAERME':
            src_vertex = next((v for v in vertices if v['id'] == edge['src']), None)
            dst_vertex = next((v for v in vertices if v['id'] == edge['dst']), None)

            if src_vertex and dst_vertex:
                waermeabnehmer.append({
                    'Lieferant': src_vertex.get('name', src_vertex['id']),
                    'Abnehmer': dst_vertex.get('name', dst_vertex['id']),
                    'Beziehung': edge.get('beziehung', 'Wärmelieferung')
                })

    return waermeabnehmer

def get_beziehungs_matrix(vertices, edges):
    """Get relationship matrix between companies"""
    firmen = {v['id']: v.get('name') for v in vertices if v.get('label') == 'Firma'}

    matrix_data = []
    for edge in edges:
        src = edge['src']
        dst = edge['dst']

        if src in firmen and dst in firmen:
            matrix_data.append({
                'Firma 1': firmen[src],
                'Firma 2': firmen[dst],
                'Beziehung': edge.get('relationship', 'UNKNOWN'),
                'Details': edge.get('rolle', '')
            })

    return pd.DataFrame(matrix_data) if matrix_data else pd.DataFrame()

def get_gesellschafter_ubersicht(vertices, edges):
    """Get overview of shareholders and participations"""
    gesellschafter = []

    for edge in edges:
        if edge.get('relationship') == 'GESELLSCHAFTER_VON':
            src_vertex = next((v for v in vertices if v['id'] == edge['src']), None)
            dst_vertex = next((v for v in vertices if v['id'] == edge['dst']), None)

            if src_vertex and dst_vertex:
                gesellschafter.append({
                    'Gesellschafter': src_vertex.get('name', src_vertex['id']),
                    'Gesellschaft': dst_vertex.get('name', dst_vertex['id']),
                    'Typ Gesellschafter': src_vertex.get('label', ''),
                    'Typ Gesellschaft': dst_vertex.get('label', '')
                })

    return pd.DataFrame(gesellschafter) if gesellschafter else pd.DataFrame()

def search_contracts(contracts, search_term):
    """Search contracts by free text across all fields"""
    search_term_lower = search_term.lower()
    results = []

    for c in contracts:
        datum = c.get('datum', {}) or {}
        finanzen = c.get('finanzen', {}) or {}
        energie = c.get('energie', {}) or {}
        searchable_fields = [
            c.get('dateiname', ''),
            c.get('vertragsart', ''),
            c.get('vertragstitel', '') or '',
            c.get('vertragsgegenstand', '') or '',
            ', '.join([p.get('name', '') for p in c.get('parteien', [])]),
            ', '.join([p.get('rolle', '') for p in c.get('parteien', []) if p.get('rolle')]),
            ', '.join(c.get('standorte', [])) if isinstance(c.get('standorte'), list) else c.get('standorte', ''),
            str(datum.get('abschluss', '') or ''),
            str(datum.get('beginn', '') or ''),
            str(datum.get('ende', '') or ''),
            c.get('laufzeit', '') or '',
            c.get('kuendigungsfrist', '') or '',
            c.get('gerichtsstand', '') or '',
            str(finanzen.get('verguetung', '') or ''),
            str(finanzen.get('vertragswert_jaehrlich', '') or ''),
            ', '.join(str(m) for m in (energie.get('malo_ids', []) or [])),
            str(energie.get('pv_leistung_kwp', '') or ''),
            c.get('vertrags_id', ''),
        ]

        if any(search_term_lower in str(field).lower() for field in searchable_fields):
            results.append(c)

    return results

def get_ego_graph(G, contract_id, depth=1):
    """Get ego graph for a specific contract"""
    if contract_id not in G:
        return nx.DiGraph()

    ego_nodes = set([contract_id])

    for _ in range(depth):
        new_nodes = set()
        for node in ego_nodes:
            new_nodes.update(G.predecessors(node))
            new_nodes.update(G.successors(node))
        ego_nodes.update(new_nodes)

    return G.subgraph(ego_nodes)

def export_to_csv(df):
    """Export dataframe to CSV"""
    return df.to_csv(index=False).encode('utf-8')

def export_to_excel(dfs_dict):
    """Export multiple dataframes to Excel"""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in dfs_dict.items():
        ws = wb.create_sheet(sheet_name[:31])

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="742774", end_color="742774", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx).value = value

        for column in ws.columns:
            max_length = 20
            column_letter = column[0].column_letter
            ws.column_dimensions[column_letter].width = min(max_length, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Predefined queries
PREDEFINED_QUERIES = {
    "── Kein Query ──": {},
    "── Vertragsarten ──": {},
    "Wärmelieferverträge in Vrees": {"vertragsart": ["Wärmeliefervertrag"], "standort": ["Vrees"]},
    "Wärmelieferverträge Alte Heide": {"vertragsart": ["Wärmeliefervertrag"], "standort_wildcard": "Alte Heide"},
    "Wärmelieferverträge Witte Moor": {"vertragsart": ["Wärmeliefervertrag"], "standort_wildcard": "Witte Moor"},
    "Betriebsführung Marka Taler": {"vertragsart": ["Betriebsführungsvertrag"], "firma": ["Marka Taler GmbH & Co. KG"]},
    "Gestattungsverträge Sonnentaler": {"vertragsart": ["Gestattungsvertrag (PV-Dachanlage)"], "projekt_wildcard": "Sonnentaler"},
    "Miet- und Nutzungsverträge": {"vertragsart": ["Mietvertrag", "Nutzungsvertrag"]},
    "── Termine & Fristen ──": {},
    "Verträge die 2026 enden": {"datum_ende_jahr": [2026]},
    "Verträge die 2025-2027 enden": {"datum_ende_jahr": [2025, 2026, 2027]},
    "Alle Nachträge": {"nachtrag": "Nur Nachträge"},
    "── Personen & Firmen ──": {},
    "Alle Verträge Niedenhof Detlef": {"person": ["Niedenhof Detlef"]},
    "Alle Verträge Niedenhof Andrea": {"person": ["Niedenhof Andrea"]},
    "── Technik ──": {},
    "Verträge > 10 Seiten": {"seiten_min": 11},
}

# Main app
def main():
    """Main Streamlit application"""

    # Load data
    vertices, edges, contracts = load_data()
    G = build_network_graph(vertices, edges)

    # ── SIDEBAR: All Filters ──
    st.sidebar.markdown("## Query-Vorlagen")
    selected_query = st.sidebar.selectbox("Vorgefertigte Query", list(PREDEFINED_QUERIES.keys()), key="sidebar_query")
    query_params = PREDEFINED_QUERIES.get(selected_query, {})

    # Apply query preset to defaults
    q_firma_default = query_params.get("firma", [])
    q_art_default = query_params.get("vertragsart", [])
    q_standort_default = query_params.get("standort", [])
    q_person_default = query_params.get("person", [])
    q_nachtrag_default = query_params.get("nachtrag")
    q_seiten_min = query_params.get("seiten_min")

    st.sidebar.markdown("---")
    st.sidebar.markdown("## Filter")
    filter_count_placeholder = st.sidebar.empty()
    reset_placeholder = st.sidebar.empty()

    # Node-Type filter (above all other filters)
    all_node_types = ['Vertrag', 'Firma', 'Person', 'Projekt', 'Anlage', 'MaLo', 'Dokument']
    default_node_types = ['Vertrag', 'Firma', 'Person', 'Projekt']
    selected_node_types = st.sidebar.multiselect(
        "Knoten",
        all_node_types,
        default=default_node_types,
        key="sb_node_types"
    )

    # Helper: get available options from a set of contracts
    def get_options_from_contracts(c_list):
        firmen = sorted(set(
            p.get('name') for c in c_list for p in c.get('parteien', []) if p.get('typ') == 'Firma'
        ))
        personen = sorted(set(
            p.get('name') for c in c_list for p in c.get('parteien', []) if p.get('typ') == 'Person'
        ))
        arten = sorted(set(c.get('vertragsart') for c in c_list if c.get('vertragsart')))
        c_ids = set(c.get('vertrags_id') for c in c_list)
        projekt_ids = set()
        for edge in edges:
            if edge['src'] in c_ids and edge.get('relationship') == 'HAT_STANDORT':
                projekt_ids.add(edge['dst'])
        projekte = sorted(
            v.get('name') for v in vertices
            if v.get('label') == 'Projekt' and v['id'] in projekt_ids
        )
        return firmen, projekte, arten, personen

    # Cascading filters in sidebar
    all_firmen = get_firmen_list(vertices)
    selected_firmen = st.sidebar.multiselect("Firma", all_firmen, default=[f for f in q_firma_default if f in all_firmen], key="sb_firmen")

    if selected_firmen:
        contracts_after_firma = filter_contracts(contracts, vertices, edges, selected_firmen, [], [], [])
    else:
        contracts_after_firma = contracts
    _, available_projekte, _, _ = get_options_from_contracts(contracts_after_firma)
    selected_projekte = st.sidebar.multiselect("Projekt", available_projekte, key="sb_projekte")

    if selected_firmen or selected_projekte:
        contracts_after_fp = filter_contracts(contracts, vertices, edges, selected_firmen, selected_projekte, [], [])
    else:
        contracts_after_fp = contracts
    _, _, available_arten, _ = get_options_from_contracts(contracts_after_fp)
    selected_arten = st.sidebar.multiselect("Vertragsart", available_arten, default=[a for a in q_art_default if a in available_arten], key="sb_arten")

    if selected_firmen or selected_projekte or selected_arten:
        contracts_after_fpa = filter_contracts(contracts, vertices, edges, selected_firmen, selected_projekte, selected_arten, [])
    else:
        contracts_after_fpa = contracts
    _, _, _, available_personen = get_options_from_contracts(contracts_after_fpa)
    selected_personen = st.sidebar.multiselect("Person", available_personen, default=[p for p in q_person_default if p in available_personen], key="sb_personen")

    # Kreuzfilterung: Erweitert-Filter basieren auf vorgefiltertem Vertragsbestand
    contracts_after_all_main = filter_contracts(
        contracts, vertices, edges,
        selected_firmen, selected_projekte, selected_arten, selected_personen
    )

    selected_standorte = []  # Standort-Filter entfernt (identisch mit Projekt)

    available_rollen = sorted(set(
        e.get('rolle', '') for e in edges
        if e.get('rolle')
        and (e['src'] in set(c.get('vertrags_id') for c in contracts_after_all_main)
             or e['dst'] in set(c.get('vertrags_id') for c in contracts_after_all_main))
    ))
    selected_rollen = st.sidebar.multiselect("Rolle", available_rollen, key="sb_rolle")

    # Datum Ende (Jahr) Filter — kreuzgefiltert
    contracts_after_extended = filter_contracts(
        contracts, vertices, edges,
        selected_firmen, selected_projekte, selected_arten, selected_personen,
        selected_standorte, selected_rollen
    )
    available_ende_jahre = sorted(set(
        c.get('datum', {}).get('ende_jahr')
        for c in contracts_after_extended
        if c.get('datum', {}).get('ende_jahr')
    ))
    q_ende_jahre = query_params.get("datum_ende_jahr", [])
    selected_ende_jahre = st.sidebar.multiselect(
        "Vertragsende (Jahr)",
        available_ende_jahre,
        default=[j for j in q_ende_jahre if j in available_ende_jahre],
        key="sb_ende_jahr"
    )

    nachtrag_options = ["Alle", "Nur Originale", "Nur Nachträge"]
    nachtrag_idx = nachtrag_options.index(q_nachtrag_default) if q_nachtrag_default in nachtrag_options else 0
    nachtrag_option = st.sidebar.radio("Nachtrag-Status", nachtrag_options, index=nachtrag_idx, key="sb_nachtrag")

    seiten_default = (q_seiten_min, 38) if q_seiten_min else (1, 38)
    seiten_range = st.sidebar.slider("Seitenzahl", 1, 38, seiten_default, key="sb_seiten")

    # Final filtered data
    sidebar_nachtrag = nachtrag_option if nachtrag_option != "Alle" else None
    sidebar_seiten = seiten_range if seiten_range != (1, 38) else None
    any_filter_active = (selected_firmen or selected_arten or selected_projekte or selected_personen
                        or selected_rollen or sidebar_nachtrag or sidebar_seiten
                        or selected_ende_jahre)

    # Projekt/Standort wildcard from query
    q_projekt_wc = query_params.get("projekt_wildcard")
    q_standort_wc = query_params.get("standort_wildcard")

    if any_filter_active or q_projekt_wc or q_standort_wc:
        filtered_contracts = filter_contracts(
            contracts, vertices, edges,
            selected_firmen, selected_projekte, selected_arten, selected_personen,
            selected_standorte, selected_rollen, sidebar_nachtrag, sidebar_seiten
        )
        # Datum Ende Jahr filter
        if selected_ende_jahre:
            filtered_contracts = [
                c for c in filtered_contracts
                if c.get('datum', {}).get('ende_jahr') in selected_ende_jahre
            ]
        if q_projekt_wc:
            projekt_nodes = [v['id'] for v in vertices
                            if v.get('label') == 'Projekt' and q_projekt_wc.lower() in v.get('name', '').lower()]
            contract_ids = set()
            for edge in edges:
                if edge['dst'] in projekt_nodes and 'vertrag:' in edge['src']:
                    contract_ids.add(edge['src'])
            filtered_contracts = [c for c in filtered_contracts if c.get('vertrags_id') in contract_ids]
        if q_standort_wc:
            filtered_contracts = [
                c for c in filtered_contracts
                if any(q_standort_wc.lower() in s.lower() for s in c.get('standorte', []) if s)
                or q_standort_wc.lower() in (c.get('dateiname', '') or '').lower()
            ]

        # Build graph from filtered contracts — ensures graph matches all filters
        filtered_contract_ids = set(c.get('vertrags_id') for c in filtered_contracts)
        relevant_nodes = set(filtered_contract_ids)
        for edge in edges:
            if edge['src'] in filtered_contract_ids:
                relevant_nodes.add(edge['dst'])
            if edge['dst'] in filtered_contract_ids:
                relevant_nodes.add(edge['src'])
        # Apply node-type filter
        if selected_node_types:
            vertex_labels = {v['id']: v.get('label') for v in vertices}
            relevant_nodes = {n for n in relevant_nodes
                             if vertex_labels.get(n) in selected_node_types}
            # Always keep contract nodes
            relevant_nodes |= filtered_contract_ids
        filtered_graph = G.subgraph(relevant_nodes)
    else:
        filtered_contracts = contracts
        # Apply node-type filter even without contract filters
        if selected_node_types and set(selected_node_types) != set(all_node_types):
            vertex_labels = {v['id']: v.get('label') for v in vertices}
            contract_ids = set(c.get('vertrags_id') for c in contracts)
            relevant_nodes = {v['id'] for v in vertices if v.get('label') in selected_node_types}
            relevant_nodes |= contract_ids
            filtered_graph = G.subgraph(relevant_nodes)
        else:
            filtered_graph = G

    # Fill the filter count placeholder at top of sidebar
    filter_count_placeholder.markdown(f"**{len(filtered_contracts)}** von {len(contracts)} Verträgen")
    if any_filter_active:
        if reset_placeholder.button("Alle Filter löschen", key="reset_filters"):
            # Multiselects auf leer setzen
            for key in ["sb_firmen", "sb_projekte", "sb_arten", "sb_personen",
                         "sb_rolle", "sb_ende_jahr"]:
                st.session_state[key] = []
            # Nachtrag auf "Alle" (Index 0)
            st.session_state["sb_nachtrag"] = "Alle"
            # Seitenzahl auf vollen Bereich
            st.session_state["sb_seiten"] = (1, 38)
            st.rerun()

    # Create tabs
    tab1, tab2, tab_query, tab3, tab4 = st.tabs([
        "Graph-Visualisierung",
        "Tabellen-Generator",
        "Query-Center",
        "Dashboard-Übersicht",
        "Detail-Ansicht"
    ])

    # ── TAB 1: GRAPH VISUALIZATION ──
    with tab1:
        # Compact legend + stats in one row
        n_knoten = len(filtered_graph.nodes())
        n_kanten = len(filtered_graph.edges())
        n_vertraege = len([v for v in filtered_graph.nodes()
                           if any(n['id'] == v and n.get('label') == 'Vertrag' for n in vertices)])
        n_firmen = len([v for v in filtered_graph.nodes()
                        if any(n['id'] == v and n.get('label') == 'Firma' for n in vertices)])

        legend_items = ""
        for lbl, color in NODE_COLORS_SIMPLE.items():
            legend_items += f"<span style='display:flex;align-items:center;gap:4px;'><span style='width:12px;height:12px;border-radius:50%;background:{color};display:inline-block;'></span>{lbl}</span>"

        stats_text = f"Knoten: <b>{n_knoten}</b> &nbsp;|&nbsp; Kanten: <b>{n_kanten}</b> &nbsp;|&nbsp; Verträge: <b>{n_vertraege}</b> &nbsp;|&nbsp; Firmen: <b>{n_firmen}</b>"

        legend_html = f"""<div style='display:flex; justify-content:space-between; align-items:center; margin:4px 0 8px 0; color:#333333; font-size:13px; font-family:Segoe UI,sans-serif;'>
            <div style='display:flex; gap:12px; flex-wrap:wrap;'>{legend_items}</div>
            <div style='white-space:nowrap;'>{stats_text}</div>
        </div>"""
        st.markdown(legend_html, unsafe_allow_html=True)

        try:
            net_graph = create_pyvis_graph(filtered_graph)
            html_string = net_graph.generate_html()
            st.components.v1.html(html_string, height=900)
        except Exception as e:
            st.error(f"Fehler bei Graphvisualisierung: {e}")

    # ── TAB 2: TABLES GENERATOR ──
    with tab2:
        col1, col2 = st.columns([2, 1])

        with col1:
            table_type = st.selectbox(
                "Tabellen-Typ",
                [
                    "Alle Verträge",
                    "Verträge pro Firma",
                    "Verträge pro Vertragsart",
                    "Verträge pro Projekt",
                    "Parteien-Übersicht",
                    "Finanz-Übersicht",
                    "Gesellschafter-Übersicht (alle)",
                    "Wärmeabnehmer-Übersicht (alle)",
                    "Beziehungs-Matrix (alle)",
                    "Custom Query"
                ]
            )

        with col2:
            st.markdown("#### Export")

        if table_type == "Alle Verträge":
            df = create_contracts_table(filtered_contracts)
            st.dataframe(df, use_container_width=True, height=900)

            col1, col2 = st.columns(2)
            with col1:
                csv_data = export_to_csv(df)
                st.download_button("CSV herunterladen", csv_data, "vertraege.csv", "text/csv")
            with col2:
                excel_data = export_to_excel({"Verträge": df})
                st.download_button("Excel herunterladen", excel_data, "vertraege.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        elif table_type == "Verträge pro Firma":
            available_firmen = sorted(set(
                p.get('name') for c in filtered_contracts for p in c.get('parteien', []) if p.get('typ') == 'Firma'
            ))
            selected_firmen_tab = st.multiselect("Firmen wählen", available_firmen, key="table_firmen_2")

            if selected_firmen_tab:
                contracts_by_firma = get_contracts_by_firma(filtered_contracts, selected_firmen_tab)

                for firma, firma_contracts in contracts_by_firma.items():
                    with st.expander(f"{firma} ({len(firma_contracts)} Verträge)"):
                        df = pd.DataFrame({
                            'Dateiname': [c.get('dateiname') for c in firma_contracts],
                            'Vertragsart': [c.get('vertragsart') for c in firma_contracts],
                            'Seiten': [c.get('seitenzahl') for c in firma_contracts],
                            'Status': ['Nachtrag' if c.get('ist_nachtrag') else 'Original' for c in firma_contracts]
                        })
                        st.dataframe(df, use_container_width=True, height=900)

        elif table_type == "Verträge pro Vertragsart":
            available_arten = sorted(set(c.get('vertragsart') for c in filtered_contracts if c.get('vertragsart')))
            selected_arten_tab = st.multiselect("Vertragsarten wählen", available_arten, key="table_arten_2")

            if selected_arten_tab:
                contracts_by_art = get_contracts_by_vertragsart(filtered_contracts, selected_arten_tab)

                for art, art_contracts in contracts_by_art.items():
                    with st.expander(f"{art} ({len(art_contracts)} Verträge)"):
                        df = pd.DataFrame({
                            'Dateiname': [c.get('dateiname') for c in art_contracts],
                            'Seiten': [c.get('seitenzahl') for c in art_contracts],
                            'Parteien': [', '.join([p.get('name') for p in c.get('parteien', [])]) for c in art_contracts]
                        })
                        st.dataframe(df, use_container_width=True, height=900)

        elif table_type == "Verträge pro Projekt":
            contracts_by_projekt = get_contracts_by_projekt(vertices, edges, filtered_contracts)

            if contracts_by_projekt:
                for projekt, projekt_contracts in contracts_by_projekt.items():
                    with st.expander(f"{projekt} ({len(projekt_contracts)} Verträge)"):
                        df = pd.DataFrame({
                            'Dateiname': [c.get('dateiname') for c in projekt_contracts],
                            'Vertragsart': [c.get('vertragsart') for c in projekt_contracts],
                            'Seiten': [c.get('seitenzahl') for c in projekt_contracts],
                            'Parteien': [', '.join([p.get('name') for p in c.get('parteien', [])]) for c in projekt_contracts]
                        })
                        st.dataframe(df, use_container_width=True, height=900)
            else:
                st.info("Keine Projekte mit Verträgen gefunden.")

        elif table_type == "Parteien-Übersicht":
            parteien_ubersicht = get_parteien_ubersicht(filtered_contracts)

            df = pd.DataFrame([
                {
                    'Partei': name,
                    'Typ': info.get('Typ'),
                    'Anzahl Verträge': info.get('Anzahl Verträge')
                }
                for name, info in parteien_ubersicht.items()
            ]).sort_values('Anzahl Verträge', ascending=False)

            st.dataframe(df, use_container_width=True, height=900)

            col1, col2 = st.columns(2)
            with col1:
                csv_data = export_to_csv(df)
                st.download_button("CSV herunterladen", csv_data, "parteien.csv", "text/csv")
            with col2:
                excel_data = export_to_excel({"Parteien": df})
                st.download_button("Excel herunterladen", excel_data, "parteien.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        elif table_type == "Finanz-Übersicht":
            df = get_finanz_ubersicht(filtered_contracts)
            st.dataframe(df, use_container_width=True, height=900)

            col1, col2 = st.columns(2)
            with col1:
                csv_data = export_to_csv(df)
                st.download_button("CSV herunterladen", csv_data, "finanzen.csv", "text/csv")
            with col2:
                excel_data = export_to_excel({"Finanzen": df})
                st.download_button("Excel herunterladen", excel_data, "finanzen.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        elif table_type == "Gesellschafter-Übersicht (alle)":
            df = get_gesellschafter_ubersicht(vertices, edges)

            if not df.empty:
                st.dataframe(df, use_container_width=True, height=900)

                col1, col2 = st.columns(2)
                with col1:
                    csv_data = export_to_csv(df)
                    st.download_button("CSV herunterladen", csv_data, "gesellschafter.csv", "text/csv")
                with col2:
                    excel_data = export_to_excel({"Gesellschafter": df})
                    st.download_button("Excel herunterladen", excel_data, "gesellschafter.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.info("Keine Gesellschafter-Beziehungen gefunden.")

        elif table_type == "Wärmeabnehmer-Übersicht (alle)":
            waermeabnehmer = get_waermeabnehmer_ubersicht(edges, vertices)

            if waermeabnehmer:
                df = pd.DataFrame(waermeabnehmer)
                st.dataframe(df, use_container_width=True, height=900)

                col1, col2 = st.columns(2)
                with col1:
                    csv_data = export_to_csv(df)
                    st.download_button("CSV herunterladen", csv_data, "waermeabnehmer.csv", "text/csv")
                with col2:
                    excel_data = export_to_excel({"Wärmeabnehmer": df})
                    st.download_button("Excel herunterladen", excel_data, "waermeabnehmer.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.info("Keine Wärmelieferungsbeziehungen gefunden.")

        elif table_type == "Beziehungs-Matrix (alle)":
            df = get_beziehungs_matrix(vertices, edges)

            if not df.empty:
                st.dataframe(df, use_container_width=True, height=900)

                col1, col2 = st.columns(2)
                with col1:
                    csv_data = export_to_csv(df)
                    st.download_button("CSV herunterladen", csv_data, "beziehungen.csv", "text/csv")
                with col2:
                    excel_data = export_to_excel({"Beziehungen": df})
                    st.download_button("Excel herunterladen", excel_data, "beziehungen.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.info("Keine Unternehmensbeziehungen gefunden.")

        elif table_type == "Custom Query":
            st.markdown("**Beispiel-Suchbegriffe:**")
            example_cols = st.columns(6)
            examples = ["Niedenhof", "Vrees", "Sonnentaler", "Wärme", "2026", "Mietvertrag"]
            search_term = ""
            for i, ex in enumerate(examples):
                with example_cols[i]:
                    if st.button(ex, key=f"ex_{ex}"):
                        search_term = ex

            user_search = st.text_input("Suchbegriff eingeben (durchsucht alle Felder)", search_term)
            if user_search:
                search_term = user_search

            if search_term:
                results = search_contracts(filtered_contracts, search_term)

                if results:
                    st.success(f"{len(results)} Verträge gefunden für '{search_term}'")
                    df = create_contracts_table(results)
                    st.dataframe(df, use_container_width=True, height=900)

                    col1, col2 = st.columns(2)
                    with col1:
                        csv_data = export_to_csv(df)
                        st.download_button("CSV", csv_data, "query_results.csv", "text/csv", key="cq_csv")
                    with col2:
                        excel_data = export_to_excel({"Query": df})
                        st.download_button("Excel", excel_data, "query_results.xlsx",
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="cq_xlsx")
                else:
                    st.info(f"Keine Verträge gefunden für '{search_term}'.")

    # ── TAB QUERY-CENTER ──
    with tab_query:
        st.markdown("### Beziehungs-Abfragen")
        st.info("Vertragsfilter laufen über die Sidebar (links). Hier findest du Spezial-Queries für Beziehungen im Graph.")

        edge_query = st.selectbox("Abfrage", [
            "── Bitte wählen ──",
            "Gesellschafterstruktur (aus Verträgen)",
            "Wer liefert Wärme an wen?",
            "Wer betreibt was? (Betriebsführung)",
            "Direktvermarktung",
            "Alle Parteien-Beziehungen",
            "Alle Standort-Beziehungen",
            "Alle Anlagen-Beziehungen",
            "Alle MaLo-Zuordnungen",
        ], key="edge_query")

        # Edge-based queries (actual relationship types in data)
        edge_type_map = {
            "Alle Parteien-Beziehungen": "PARTEI_VON",
            "Alle Standort-Beziehungen": "HAT_STANDORT",
            "Alle Anlagen-Beziehungen": "BETRIFFT_ANLAGE",
            "Alle MaLo-Zuordnungen": "HAT_MALO",
        }

        if edge_query in edge_type_map:
            edge_type = edge_type_map[edge_query]
            results = []
            for edge in edges:
                if edge.get('relationship') == edge_type:
                    src_v = next((v for v in vertices if v['id'] == edge['src']), None)
                    dst_v = next((v for v in vertices if v['id'] == edge['dst']), None)
                    if src_v and dst_v:
                        row = {
                            'Von': src_v.get('name', edge['src']),
                            'Typ (Von)': src_v.get('label', ''),
                            'An': dst_v.get('name', edge['dst']),
                            'Typ (An)': dst_v.get('label', ''),
                        }
                        if edge.get('rolle'):
                            row['Rolle'] = edge['rolle']
                        results.append(row)
            if results:
                st.success(f"{len(results)} Beziehungen gefunden")
                st.dataframe(pd.DataFrame(results), use_container_width=True, height=900)
            else:
                st.info("Keine Ergebnisse gefunden.")

        # Contract-type-based queries (derive from vertragsart + parteien)
        elif edge_query == "Gesellschafterstruktur (aus Verträgen)":
            results = []
            for c in filtered_contracts:
                if 'Gesellschaft' in (c.get('vertragsart') or '').lower() or 'gesellschaft' in (c.get('vertragsart') or '').lower():
                    parteien = c.get('parteien', [])
                    for p in parteien:
                        results.append({
                            'Gesellschaft/Vertrag': c.get('vertragstitel') or c.get('dateiname', ''),
                            'Partei': p.get('name', ''),
                            'Typ': p.get('typ', ''),
                            'Rolle': p.get('rolle', ''),
                            'Vertragsart': c.get('vertragsart', ''),
                            'Standort': ', '.join(c.get('standorte', [])) if isinstance(c.get('standorte'), list) else (c.get('standorte') or ''),
                        })
            if results:
                st.success(f"{len(results)} Gesellschafter-Beziehungen gefunden")
                st.dataframe(pd.DataFrame(results), use_container_width=True, height=900)
            else:
                st.info("Keine Gesellschafterverträge in der aktuellen Filterung gefunden.")

        elif edge_query == "Wer liefert Wärme an wen?":
            results = []
            for c in filtered_contracts:
                if 'wärme' in (c.get('vertragsart') or '').lower() or 'Wärme' in (c.get('vertragsart') or ''):
                    parteien = c.get('parteien', [])
                    lieferant = next((p for p in parteien if 'liefer' in (p.get('rolle') or '').lower()), None)
                    kunde = next((p for p in parteien if 'kunde' in (p.get('rolle') or '').lower() or 'abnehm' in (p.get('rolle') or '').lower() or 'nutz' in (p.get('rolle') or '').lower()), None)
                    results.append({
                        'Lieferant': lieferant.get('name') if lieferant else (parteien[0].get('name') if parteien else 'N/A'),
                        'Kunde/Abnehmer': kunde.get('name') if kunde else (parteien[1].get('name') if len(parteien) > 1 else 'N/A'),
                        'Vertrag': c.get('dateiname', ''),
                        'Standort': ', '.join(c.get('standorte', [])) if isinstance(c.get('standorte'), list) else (c.get('standorte') or ''),
                    })
            if results:
                st.success(f"{len(results)} Wärmelieferbeziehungen gefunden")
                st.dataframe(pd.DataFrame(results), use_container_width=True, height=900)
            else:
                st.info("Keine Wärmelieferverträge gefunden.")

        elif edge_query == "Wer betreibt was? (Betriebsführung)":
            results = []
            for c in filtered_contracts:
                if 'betriebsführ' in (c.get('vertragsart') or '').lower() or 'betriebsführ' in (c.get('vertragstitel') or '').lower():
                    parteien = c.get('parteien', [])
                    results.append({
                        'Auftraggeber': next((p.get('name') for p in parteien if 'auftragge' in (p.get('rolle') or '').lower()), parteien[0].get('name') if parteien else 'N/A'),
                        'Betriebsführer': next((p.get('name') for p in parteien if 'auftragnehm' in (p.get('rolle') or '').lower() or 'betrieb' in (p.get('rolle') or '').lower()), parteien[1].get('name') if len(parteien) > 1 else 'N/A'),
                        'Vertrag': c.get('dateiname', ''),
                        'Standort': ', '.join(c.get('standorte', [])) if isinstance(c.get('standorte'), list) else (c.get('standorte') or ''),
                    })
            if results:
                st.success(f"{len(results)} Betriebsführungsverträge gefunden")
                st.dataframe(pd.DataFrame(results), use_container_width=True, height=900)
            else:
                st.info("Keine Betriebsführungsverträge gefunden.")

        elif edge_query == "Direktvermarktung":
            results = []
            for c in filtered_contracts:
                if 'direktvermarkt' in (c.get('vertragsart') or '').lower() or 'direktvermarkt' in (c.get('vertragstitel') or '').lower():
                    parteien = c.get('parteien', [])
                    results.append({
                        'Vermarkter': next((p.get('name') for p in parteien if 'vermarkt' in (p.get('rolle') or '').lower()), parteien[0].get('name') if parteien else 'N/A'),
                        'Erzeuger': next((p.get('name') for p in parteien if 'erzeug' in (p.get('rolle') or '').lower()), parteien[1].get('name') if len(parteien) > 1 else 'N/A'),
                        'Vertrag': c.get('dateiname', ''),
                        'Standort': ', '.join(c.get('standorte', [])) if isinstance(c.get('standorte'), list) else (c.get('standorte') or ''),
                    })
            if results:
                st.success(f"{len(results)} Direktvermarktungsverträge gefunden")
                st.dataframe(pd.DataFrame(results), use_container_width=True, height=900)
            else:
                st.info("Keine Direktvermarktungsverträge gefunden.")

    # ── TAB 3: DASHBOARD ──
    with tab3:
        # KPIs based on filtered data
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Verträge", len(filtered_contracts))

        with col2:
            firmen_in_filtered = set(
                p.get('name') for c in filtered_contracts for p in c.get('parteien', []) if p.get('typ') == 'Firma'
            )
            st.metric("Firmen", len(firmen_in_filtered))

        with col3:
            # Count projects linked to filtered contracts
            filtered_ids = set(c.get('vertrags_id') for c in filtered_contracts)
            projekt_set = set()
            for edge in edges:
                if edge['src'] in filtered_ids and edge.get('relationship') == 'HAT_STANDORT':
                    projekt_set.add(edge['dst'])
            st.metric("Projekte", len(projekt_set))

        with col4:
            personen_in_filtered = set(
                p.get('name') for c in filtered_contracts for p in c.get('parteien', []) if p.get('typ') == 'Person'
            )
            st.metric("Personen", len(personen_in_filtered))

        # Charts
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("#### Verträge pro Vertragsart")
            vertragsart_counts = {}
            for c in filtered_contracts:
                art = c.get('vertragsart', 'Unbekannt')
                vertragsart_counts[art] = vertragsart_counts.get(art, 0) + 1

            if vertragsart_counts:
                terp_palette = ['#742774', '#80C6FF', '#CAEEE9', '#D8D8EE', '#FFEBC0',
                                '#D4E7F6', '#E6E0F0', '#F5CCCF', '#CAF0CC', '#C3CAF9',
                                '#E5D2E3', '#C3F8F9', '#FED5D1', '#FABBC3']
                fig = go.Figure(data=[
                    go.Pie(labels=list(vertragsart_counts.keys()),
                          values=list(vertragsart_counts.values()),
                          hole=0.3,
                          marker=dict(colors=terp_palette))
                ])
                fig.update_layout(height=450, showlegend=True, paper_bgcolor='#F8F8F8', plot_bgcolor='#FFFFFF', font_color='#333333')
                st.plotly_chart(fig, use_container_width=True)

        with col2:
            st.markdown("#### Vertragsstatistiken")
            nachtrag_count = len([c for c in filtered_contracts if c.get('ist_nachtrag')])
            original_count = len(filtered_contracts) - nachtrag_count

            stats_data = {
                'Status': ['Originale', 'Nachträge'],
                'Anzahl': [original_count, nachtrag_count]
            }

            fig = go.Figure(data=[
                go.Bar(x=stats_data['Status'], y=stats_data['Anzahl'], marker_color=['#80C6FF', '#F5CCCF'])
            ])
            fig.update_layout(height=450, showlegend=False, paper_bgcolor='#F8F8F8', plot_bgcolor='#FFFFFF', font_color='#333333')
            st.plotly_chart(fig, use_container_width=True)

        # Projekte chart
        st.markdown("#### Verträge pro Projekt")
        contracts_by_projekt = get_contracts_by_projekt(vertices, edges, filtered_contracts)

        if contracts_by_projekt:
            projekt_names = list(contracts_by_projekt.keys())
            projekt_counts = [len(c) for c in contracts_by_projekt.values()]

            fig = go.Figure(data=[
                go.Bar(x=projekt_names, y=projekt_counts, marker_color='#742774')
            ])
            fig.update_layout(height=450, xaxis_title="Projekt", yaxis_title="Anzahl Verträge", paper_bgcolor='#F8F8F8', plot_bgcolor='#FFFFFF', font_color='#333333')
            st.plotly_chart(fig, use_container_width=True)

    # ── TAB 4: DETAIL VIEW ──
    with tab4:
        contract_names = [c.get('dateiname', c.get('vertrags_id')) for c in filtered_contracts]
        selected_contract_name = st.selectbox("Vertrag wählen", contract_names)

        if selected_contract_name:
            selected_contract = next((c for c in filtered_contracts
                                     if c.get('dateiname') == selected_contract_name), None)

            if selected_contract:
                # 50/50 Layout: Ego-Graph links, Details rechts
                graph_col, detail_col = st.columns(2)

                # --- LEFT: Ego Graph ---
                with graph_col:
                    # Legende über dem Graph
                    legend_items = ""
                    for lbl, color in NODE_COLORS_SIMPLE.items():
                        legend_items += f"<span style='display:flex;align-items:center;gap:4px;'><span style='width:12px;height:12px;border-radius:50%;background:{color};display:inline-block;'></span>{lbl}</span>"
                    st.markdown(f"<div style='display:flex; gap:12px; flex-wrap:wrap; margin:4px 0 8px 0; color:#333333; font-size:13px; font-family:Segoe UI,sans-serif;'>{legend_items}</div>", unsafe_allow_html=True)

                    contract_id = selected_contract.get('vertrags_id')
                    if contract_id in G:
                        ego = get_ego_graph(G, contract_id)
                        if ego.number_of_nodes() > 0:
                            try:
                                ego_vis = create_ego_pyvis_graph(ego)
                                html_string = ego_vis.generate_html()
                                st.components.v1.html(html_string, height=700)
                            except Exception as e:
                                st.error(f"Fehler bei Ego-Graph: {e}")
                        else:
                            st.info("Keine direkten Verbindungen gefunden.")

                # --- RIGHT: All details compact ---
                with detail_col:
                    datum = selected_contract.get('datum', {}) or {}
                    standorte = selected_contract.get('standorte', [])
                    if isinstance(standorte, list):
                        standorte = ', '.join(s for s in standorte if s)
                    finanzen = selected_contract.get('finanzen', {}) or {}
                    energie = selected_contract.get('energie', {}) or {}

                    # Parteien
                    parteien_html = ""
                    for party in selected_contract.get('parteien', []):
                        rolle = f" – {party.get('rolle')}" if party.get('rolle') else ''
                        parteien_html += f"<div>{party.get('name')} ({party.get('typ')}{rolle})</div>"
                    if not parteien_html:
                        parteien_html = "<div>N/A</div>"

                    # Finanzen
                    fin_lines = []
                    if finanzen.get('verguetung'):
                        fin_lines.append(f"Vergütung: {finanzen['verguetung']}")
                    if finanzen.get('verguetungsbasis'):
                        fin_lines.append(f"Basis: {finanzen['verguetungsbasis']}")
                    if finanzen.get('vertragswert_jaehrlich'):
                        fin_lines.append(f"Wert p.a.: {finanzen['vertragswert_jaehrlich']}")
                    if finanzen.get('arbeitspreis_ct_kwh'):
                        fin_lines.append(f"Arbeitspreis: {finanzen['arbeitspreis_ct_kwh']} ct/kWh")
                    if finanzen.get('grundpreis_eur_monat'):
                        fin_lines.append(f"Grundpreis: {finanzen['grundpreis_eur_monat']} EUR/Mo")
                    if finanzen.get('kaltmiete_eur'):
                        fin_lines.append(f"Kaltmiete: {finanzen['kaltmiete_eur']} EUR")
                    if finanzen.get('zahlungsrhythmus'):
                        fin_lines.append(f"Zahlung: {finanzen['zahlungsrhythmus']}")
                    fin_html = "<br>".join(fin_lines) if fin_lines else "<em>Keine Finanzdaten</em>"

                    # Energie
                    ene_lines = []
                    if energie.get('pv_leistung_kwp'):
                        ene_lines.append(f"PV: {energie['pv_leistung_kwp']} kWp")
                    if energie.get('anschlussleistung_kw'):
                        ene_lines.append(f"Anschluss: {energie['anschlussleistung_kw']} kW")
                    malo_ids = energie.get('malo_ids', [])
                    if malo_ids:
                        ene_lines.append(f"MaLo: {', '.join(str(m) for m in malo_ids)}")
                    ene_html = "<br>".join(ene_lines) if ene_lines else "<em>Keine Energiedaten</em>"

                    # Render all as compact HTML
                    detail_html = f"""
                    <div style="font-family:'Segoe UI',sans-serif; font-size:13px; color:#333; line-height:1.5;">
                        <div style="font-size:14px; font-weight:700; color:#742774; margin-bottom:4px;">
                            {selected_contract.get('vertragstitel') or selected_contract.get('vertragsart', '')}
                        </div>

                        <div style="display:grid; grid-template-columns:1fr 1fr; gap:2px 16px; margin-bottom:6px;">
                            <div><b>Art:</b> {selected_contract.get('vertragsart', 'N/A')}</div>
                            <div><b>Abschluss:</b> {datum.get('abschluss') or 'N/A'}</div>
                            <div><b>Status:</b> {'Nachtrag' if selected_contract.get('ist_nachtrag') else 'Original'}</div>
                            <div><b>Beginn:</b> {datum.get('beginn') or 'N/A'}</div>
                            <div><b>Seiten:</b> {selected_contract.get('seitenzahl', 'N/A')}</div>
                            <div><b>Ende:</b> {datum.get('ende') or 'N/A'}</div>
                            <div><b>Standort:</b> {standorte or 'N/A'}</div>
                            <div><b>Laufzeit:</b> {selected_contract.get('laufzeit') or 'N/A'}</div>
                        </div>

                        <hr style="margin:4px 0; border:0; border-top:1px solid #D0D0D0;">
                        <div style="font-size:13px; font-weight:600; margin:4px 0 2px;">Parteien</div>
                        {parteien_html}

                        <hr style="margin:4px 0; border:0; border-top:1px solid #D0D0D0;">
                        <div style="font-size:13px; font-weight:600; margin:4px 0 2px;">Kündigung</div>
                        <div><b>Frist:</b> {selected_contract.get('kuendigungsfrist') or 'N/A'}</div>
                        <div><b>Form:</b> {selected_contract.get('kuendigung_form') or 'N/A'}</div>
                        <div><b>Außerord.:</b> {selected_contract.get('kuendigung_ausserordentlich') or 'N/A'}</div>

                        <hr style="margin:4px 0; border:0; border-top:1px solid #D0D0D0;">
                        <div style="font-size:13px; font-weight:600; margin:4px 0 2px;">Finanzen</div>
                        {fin_html}

                        <hr style="margin:4px 0; border:0; border-top:1px solid #D0D0D0;">
                        <div style="font-size:13px; font-weight:600; margin:4px 0 2px;">Energie / Technik</div>
                        {ene_html}

                        {"<hr style='margin:4px 0; border:0; border-top:1px solid #D0D0D0;'><div style='font-size:13px; font-weight:600; margin:4px 0 2px;'>Vertragsgegenstand</div><div>" + selected_contract['vertragsgegenstand'] + "</div>" if selected_contract.get('vertragsgegenstand') else ""}
                    </div>
                    """
                    st.components.v1.html(detail_html, height=600, scrolling=True)

                # Text preview (full width below)
                if selected_contract.get('text_preview') and selected_contract.get('text_preview') != "[KEIN TEXT]":
                    st.markdown("##### Textvorschau")
                    st.text_area("Vorschau", selected_contract.get('text_preview', ''), height=200, disabled=True)

    # Multiselect tag coloring — via st.components.v1.html (allows JS)
    st.components.v1.html("""
    <script>
    const TAG_COLORS = {
        'Vertrag':  {bg: '#FABBC3', border: '#E8949F'},
        'Firma':    {bg: '#CAEEE9', border: '#9DD5CC'},
        'Person':   {bg: '#FFEBC0', border: '#E8CF97'},
        'Anlage':   {bg: '#D4E7F6', border: '#A8CCE6'},
        'Projekt':  {bg: '#D8D8EE', border: '#B3B3D6'},
        'Dokument': {bg: '#E5E5E5', border: '#C0C0C0'},
        'MaLo':     {bg: '#E6E0F0', border: '#C5BAD9'}
    };
    const LABEL_COLORS = {
        'Firma':       {bg: '#CAEEE9', border: '#9DD5CC'},
        'Projekt':     {bg: '#D8D8EE', border: '#B3B3D6'},
        'Person':      {bg: '#FFEBC0', border: '#E8CF97'},
        'Standort':    {bg: '#D4E7F6', border: '#A8CCE6'},
        'Vertragsart': {bg: '#FABBC3', border: '#E8949F'}
    };

    function colorTags() {
        const doc = window.parent.document;

        // Color tags by their text content (for Knoten-Typen)
        doc.querySelectorAll('[data-baseweb="tag"]').forEach(tag => {
            const spans = tag.querySelectorAll('span');
            for (const span of spans) {
                const text = span.textContent.trim();
                if (TAG_COLORS[text]) {
                    tag.style.setProperty('background-color', TAG_COLORS[text].bg, 'important');
                    tag.style.setProperty('border-color', TAG_COLORS[text].border, 'important');
                    tag.querySelectorAll('span').forEach(s => s.style.setProperty('color', '#333', 'important'));
                    const actionBtn = tag.querySelector('[data-baseweb="tag-action"]');
                    if (actionBtn) actionBtn.style.setProperty('color', '#555', 'important');
                    break;
                }
            }
        });

        // Color all tags under specific labeled multiselects
        doc.querySelectorAll('label').forEach(label => {
            const labelText = label.textContent.trim();
            if (LABEL_COLORS[labelText]) {
                const widget = label.closest('[data-testid="stWidgetLabel"]');
                if (!widget) return;
                const container = widget.parentElement;
                if (!container) return;
                container.querySelectorAll('[data-baseweb="tag"]').forEach(tag => {
                    tag.style.setProperty('background-color', LABEL_COLORS[labelText].bg, 'important');
                    tag.style.setProperty('border-color', LABEL_COLORS[labelText].border, 'important');
                    tag.querySelectorAll('span').forEach(s => s.style.setProperty('color', '#333', 'important'));
                    const actionBtn = tag.querySelector('[data-baseweb="tag-action"]');
                    if (actionBtn) actionBtn.style.setProperty('color', '#555', 'important');
                });
            }
        });
    }

    // Translate placeholder text — only target actual placeholder elements (leaf nodes)
    function translatePlaceholders() {
        const doc = window.parent.document;
        doc.querySelectorAll('[data-baseweb="select"] [data-baseweb="select-placeholder"], [data-baseweb="select"] .st-emotion-cache-placeholder').forEach(el => {
            if (el.textContent.trim() === 'Choose an option') {
                el.textContent = 'Bitte auswählen...';
            }
        });
    }

    // Run with guard to prevent feedback loops from MutationObserver
    let running = false;
    function runAll() {
        if (running) return;
        running = true;
        try { colorTags(); translatePlaceholders(); }
        finally { setTimeout(() => { running = false; }, 100); }
    }
    const observer = new MutationObserver(() => setTimeout(runAll, 150));
    observer.observe(window.parent.document.body, {childList: true, subtree: true});
    setInterval(runAll, 1000);
    runAll();
    </script>
    """, height=0)

    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: #999999; font-size: 11px; font-family: Segoe UI, sans-serif;'>
        TERP Vertragsmanagement | Powered by Streamlit
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
